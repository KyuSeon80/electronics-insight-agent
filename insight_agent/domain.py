"""데이터 접근 계층.

dataset_1의 스타 스키마(차원 2 + 팩트 4)를 읽어, 품질/생산/시장/통합 에이전트가
공통으로 쓰는 조회·조인 함수를 제공한다. MCP 서버(mymcp/server.py)는 이 모듈의
함수만 호출한다 — 데이터 로직과 프로토콜 계층을 분리해두면 나중에 실제 MCP SDK로
전송 계층만 바꿔도 이 파일은 그대로 재사용할 수 있다.
"""
from __future__ import annotations

import json
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

import numpy as np
import pandas as pd

from insight_agent.config import DATASET_DIR, DEFECT_SEVERITY_WEIGHTS, MIN_MONTHS_FOR_CORRELATION

TABLE_FILES = {
    "dim_company_product": "dim_company_product.csv",
    "dim_equipment": "dim_equipment.csv",
    "fact_production_run": "fact_production_run.csv",
    "fact_equipment_sensor": "fact_equipment_sensor.csv",
    "fact_quality_defect": "fact_quality_defect.csv",
    "fact_market_sales": "fact_market_sales.csv",
}


@dataclass
class Tables:
    dim_company_product: pd.DataFrame
    dim_equipment: pd.DataFrame
    fact_production_run: pd.DataFrame
    fact_equipment_sensor: pd.DataFrame
    fact_quality_defect: pd.DataFrame
    fact_market_sales: pd.DataFrame


def load_tables(data_dir: Path | str = DATASET_DIR) -> Tables:
    data_dir = Path(data_dir)
    frames = {
        name: _read_table(data_dir, filename) for name, filename in TABLE_FILES.items()
    }
    return Tables(**frames)


def _read_table(data_dir: Path, csv_filename: str) -> pd.DataFrame:
    """csv_filename(예: dim_equipment.csv)을 우선 찾고, 없으면 같은 이름의
    .xlsx를 읽는다.

    dataset_1은 CSV 또는 xlsx로 배포될 수 있다 -- 개별 xlsx는 CSV를 단일
    시트로 그대로 감싼 형태라 pd.read_excel로 읽어도 컬럼 스키마는 동일하다.
    """
    csv_path = data_dir / csv_filename
    if csv_path.exists():
        return pd.read_csv(csv_path, encoding="utf-8-sig")
    xlsx_path = csv_path.with_suffix(".xlsx")
    if xlsx_path.exists():
        return pd.read_excel(xlsx_path)
    raise FileNotFoundError(f"{csv_path} 또는 {xlsx_path}를 찾을 수 없음")


def df_to_records(df: pd.DataFrame) -> list[dict]:
    """DataFrame -> JSON-safe list[dict] (numpy int64/float64를 안전하게 변환)."""
    if df.empty:
        return []
    return json.loads(df.to_json(orient="records", force_ascii=False))


def get_production_anomalies(
    tables: Tables,
    factory_code: Optional[str] = None,
    min_oee_pct: float = 90.0,
) -> pd.DataFrame:
    """OEE가 임계치 미만이거나 설비 센서 anomaly_flag가 발생한 생산 실행 목록."""
    runs = tables.fact_production_run
    sensor_anomaly_runs = set(
        tables.fact_equipment_sensor.loc[
            tables.fact_equipment_sensor["anomaly_flag"] == 1, "run_id"
        ]
    )
    mask = (runs["oee_pct"] < min_oee_pct) | runs["run_id"].isin(sensor_anomaly_runs)
    if factory_code:
        mask &= runs["factory_code"] == factory_code
    result = runs.loc[mask].copy()
    result["has_sensor_anomaly"] = result["run_id"].isin(sensor_anomaly_runs)
    return result.sort_values("oee_pct")


def get_quality_defects(
    tables: Tables,
    severity: Optional[str] = None,
    category: Optional[str] = None,
) -> pd.DataFrame:
    """불량 이력에 제품 카테고리를 조인해 필터링."""
    defects = tables.fact_quality_defect.merge(
        tables.dim_company_product[["product_id", "category", "company_name"]],
        on="product_id",
        how="left",
    )
    if severity:
        defects = defects[defects["severity"] == severity]
    if category:
        defects = defects[defects["category"] == category]
    return defects.sort_values("inspection_datetime")


def get_market_impact(
    tables: Tables,
    product_id: str,
    region: Optional[str] = None,
) -> pd.DataFrame:
    """제품 하나의 매출/점유율 추이."""
    sales = tables.fact_market_sales[tables.fact_market_sales["product_id"] == product_id]
    if region:
        sales = sales[sales["region"] == region]
    return sales.sort_values("sales_month")


def build_causal_report(tables: Tables, product_id: str) -> dict:
    """생산 이상 -> 품질 불량 -> 시장 성과를 product_id 기준으로 엮은 통합 리포트.

    이 함수가 이 프로젝트의 핵심 가치다: 4개 팩트 테이블이 product_id/run_id로
    이어져 있어야만 "이 불량이 저 매출 하락의 원인인가"를 한 번에 답할 수 있다.
    """
    product_rows = tables.dim_company_product[tables.dim_company_product["product_id"] == product_id]
    if product_rows.empty:
        raise ValueError(f"unknown product_id: {product_id}")
    product = product_rows.iloc[0]

    runs = tables.fact_production_run[tables.fact_production_run["product_id"] == product_id]

    anomalies = get_production_anomalies(tables)
    anomalies = anomalies[anomalies["product_id"] == product_id]

    defects = tables.fact_quality_defect[tables.fact_quality_defect["product_id"] == product_id]
    critical_defects = defects[defects["severity"] == "Critical"]

    market = get_market_impact(tables, product_id)
    market_share_trend = market[["sales_month", "region", "market_share_est_pct", "revenue_krw"]]

    critical_defect_types = (
        json.loads(critical_defects["defect_type"].value_counts().to_json(force_ascii=False))
        if not critical_defects.empty
        else {}
    )

    return {
        "product_id": product_id,
        "model_name": product["model_name"],
        "category": product["category"],
        "company_name": product["company_name"],
        "production_run_count": int(len(runs)),
        "anomaly_run_count": int(len(anomalies)),
        "defect_count": int(len(defects)),
        "critical_defect_count": int(len(critical_defects)),
        "critical_defect_types": critical_defect_types,
        "market_share_trend": df_to_records(market_share_trend),
        "latest_market_share_pct": (
            float(market_share_trend.iloc[-1]["market_share_est_pct"])
            if not market_share_trend.empty
            else None
        ),
    }


def check_source_consistency(tables: Tables, xlsx_path: Path) -> list[dict]:
    """CSV로 읽은 각 테이블이 번들 xlsx 시트와 행수가 일치하는지 검증하는 하네스 가드레일.

    dataset_1은 현재 CSV와 xlsx가 완전히 일치하지만, 실제 운영 데이터에서는
    소스가 갈라지는 게 흔하다 -- 이 검사가 실패하면 HITL로 넘겨야 할 신호다.
    """
    import openpyxl

    mismatches: list[dict] = []
    if not Path(xlsx_path).exists():
        return [{"table": "*", "issue": "xlsx_not_found", "path": str(xlsx_path)}]

    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    for name in TABLE_FILES:
        if name not in wb.sheetnames:
            mismatches.append({"table": name, "issue": "sheet_missing"})
            continue
        ws = wb[name]
        xlsx_row_count = sum(1 for _ in ws.iter_rows(min_row=2))
        csv_row_count = len(getattr(tables, name))
        if xlsx_row_count != csv_row_count:
            mismatches.append({
                "table": name,
                "issue": "row_count_mismatch",
                "csv_rows": csv_row_count,
                "xlsx_rows": xlsx_row_count,
            })
    return mismatches


# --- 우선조치 대시보드 (docs/prd.md) --------------------------------------
# product_id/equipment_id/factory_code 축의 불량률·가중 불량 지수, 그리고
# 월별 생산-매출 조인. 컬럼명/조인 키를 아는 유일한 곳이라는 원칙을 그대로
# 지킨다 -- agents/mymcp는 이 함수들만 호출한다.


def _defect_rate_by_group(runs: pd.DataFrame, group_col: str) -> pd.DataFrame:
    """수량가중 불량률(%) = SUM(defect_qty) / SUM(actual_qty) * 100.

    run 단위 defect_rate_pct를 그대로 평균 내지 않는다 -- run마다 actual_qty가
    달라 단순평균은 소량 배치의 극단값에 왜곡되기 쉽다.
    """
    grouped = runs.groupby(group_col).agg(
        run_count=("run_id", "count"),
        actual_qty=("actual_qty", "sum"),
        good_qty=("good_qty", "sum"),
        defect_qty=("defect_qty", "sum"),
    ).reset_index()
    grouped["defect_rate_pct"] = (grouped["defect_qty"] / grouped["actual_qty"] * 100).round(2)
    return grouped.sort_values("defect_rate_pct", ascending=False).reset_index(drop=True)


def get_defect_rate_by_product(tables: Tables) -> pd.DataFrame:
    """product_id별 수량가중 불량률."""
    return _defect_rate_by_group(tables.fact_production_run, "product_id")


def get_defect_rate_by_equipment(tables: Tables) -> pd.DataFrame:
    """equipment_id별 수량가중 불량률."""
    return _defect_rate_by_group(tables.fact_production_run, "equipment_id")


def get_defect_rate_by_factory(tables: Tables) -> pd.DataFrame:
    """factory_code별 수량가중 불량률."""
    return _defect_rate_by_group(tables.fact_production_run, "factory_code")


def get_weighted_defect_index(
    tables: Tables,
    group_col: str,
    weights: Optional[dict] = None,
) -> pd.DataFrame:
    """severity(Critical/Major/Minor) 가중치를 곱한 보조 불량 지표.

    group_col은 "product_id", "equipment_id", "factory_code" 중 하나다.
    factory_code는 fact_quality_defect에 직접 없어 dim_equipment을 거쳐
    equipment_id -> factory_code로 붙인다. weights 기본값은
    config.DEFECT_SEVERITY_WEIGHTS(docs/prd.md 기준 초안, 미확정)이다.
    """
    weights = weights or DEFECT_SEVERITY_WEIGHTS
    defects = tables.fact_quality_defect.copy()
    if group_col == "factory_code":
        defects = defects.merge(
            tables.dim_equipment[["equipment_id", "factory_code"]],
            on="equipment_id",
            how="left",
        )
    defects["severity_weight"] = defects["severity"].map(weights).fillna(1.0)
    defects["weighted_defect_qty"] = defects["defect_qty"] * defects["severity_weight"]
    grouped = defects.groupby(group_col).agg(
        defect_count=("defect_id", "count"),
        defect_qty=("defect_qty", "sum"),
        critical_defect_count=("severity", lambda s: int((s == "Critical").sum())),
        weighted_defect_index=("weighted_defect_qty", "sum"),
    ).reset_index()
    return grouped.sort_values("weighted_defect_index", ascending=False).reset_index(drop=True)


def get_monthly_defect_vs_sales(
    tables: Tables,
    product_id: Optional[str] = None,
) -> pd.DataFrame:
    """product_id x 월 단위로 생산 불량률과 시장 성과를 나란히 놓은 테이블.

    fact_production_run은 run_date(일 단위), fact_market_sales는 sales_month
    (월 단위)로 그래뉼래러티가 다르다 -- run_date를 월 시작일로 맞춰 조인한다.
    factory_code/region을 잇는 키가 데이터셋에 없어(docs/prd.md 7절 UNKNOWN)
    이 조인은 product_id 단위로만 가능하다.
    """
    runs = tables.fact_production_run.copy()
    runs["sales_month"] = pd.to_datetime(runs["run_date"]).values.astype("datetime64[M]")
    monthly_production = runs.groupby(["product_id", "sales_month"]).agg(
        actual_qty=("actual_qty", "sum"),
        defect_qty=("defect_qty", "sum"),
    ).reset_index()
    monthly_production["defect_rate_pct"] = (
        monthly_production["defect_qty"] / monthly_production["actual_qty"] * 100
    ).round(2)

    sales = tables.fact_market_sales.copy()
    sales["sales_month"] = pd.to_datetime(sales["sales_month"]).values.astype("datetime64[M]")
    monthly_sales = sales.groupby(["product_id", "sales_month"]).agg(
        revenue_krw=("revenue_krw", "sum"),
        operating_profit_krw=("operating_profit_krw", "sum"),
        market_share_est_pct=("market_share_est_pct", "mean"),
    ).reset_index()
    monthly_sales["operating_margin_pct"] = (
        monthly_sales["operating_profit_krw"] / monthly_sales["revenue_krw"] * 100
    ).round(2)

    merged = monthly_production.merge(monthly_sales, on=["product_id", "sales_month"], how="inner")
    if product_id:
        merged = merged[merged["product_id"] == product_id]
    merged = merged.sort_values(["product_id", "sales_month"]).reset_index(drop=True)
    # sales_month을 "YYYY-MM" 문자열로 되돌린다 -- Timestamp를 그대로 두면
    # df_to_records()의 JSON 직렬화가 epoch 밀리초로 바꿔버려 MCP/FE
    # 소비자가 날짜를 알아볼 수 없게 된다.
    merged["sales_month"] = merged["sales_month"].dt.strftime("%Y-%m")
    return merged


def get_monthly_defect_vs_sales_by_group(
    tables: Tables,
    group_col: str,
    group_value: str,
) -> pd.DataFrame:
    """product_id뿐 아니라 equipment_id/factory_code 단위로도 월별
    불량률-매출 추이를 볼 수 있게 일반화한 버전 (FE 드릴다운/what-if용).

    group_col="product_id"면 get_monthly_defect_vs_sales와 동일하다.
    equipment_id/factory_code는 fact_market_sales와 직접 연결되지 않으므로,
    이 설비/공장을 거친 product_id 집합을 구한 뒤 그 제품들의 월별 매출
    합계를 "이 설비/공장과 연관된 매출"로 근사한다. 특정 월에 실제로 그
    설비에서 그 제품을 생산했는지는 따지지 않는다 -- get_market_impact_score_by_equipment/
    factory의 롤업(docs/prd.md 4.2절)과 같은 단순화이며, equipment_id 불량률
    자체는 fact_production_run에서 정확히 집계한다.
    """
    if group_col == "product_id":
        return get_monthly_defect_vs_sales(tables, product_id=group_value)

    if group_col not in ("equipment_id", "factory_code"):
        raise ValueError(f"group_col must be one of product_id/equipment_id/factory_code, got {group_col!r}")

    runs = tables.fact_production_run
    matched_runs = runs[runs[group_col] == group_value].copy()
    columns = [
        group_col, "sales_month", "actual_qty", "defect_qty", "defect_rate_pct",
        "revenue_krw", "operating_profit_krw", "market_share_est_pct", "operating_margin_pct",
    ]
    if matched_runs.empty:
        return pd.DataFrame(columns=columns)

    matched_runs["sales_month"] = pd.to_datetime(matched_runs["run_date"]).values.astype("datetime64[M]")
    monthly_defect = matched_runs.groupby("sales_month").agg(
        actual_qty=("actual_qty", "sum"),
        defect_qty=("defect_qty", "sum"),
    ).reset_index()
    monthly_defect["defect_rate_pct"] = (
        monthly_defect["defect_qty"] / monthly_defect["actual_qty"] * 100
    ).round(2)

    related_product_ids = matched_runs["product_id"].unique()
    sales = tables.fact_market_sales[tables.fact_market_sales["product_id"].isin(related_product_ids)].copy()
    sales["sales_month"] = pd.to_datetime(sales["sales_month"]).values.astype("datetime64[M]")
    monthly_sales = sales.groupby("sales_month").agg(
        revenue_krw=("revenue_krw", "sum"),
        operating_profit_krw=("operating_profit_krw", "sum"),
        market_share_est_pct=("market_share_est_pct", "mean"),
    ).reset_index()
    monthly_sales["operating_margin_pct"] = (
        monthly_sales["operating_profit_krw"] / monthly_sales["revenue_krw"] * 100
    ).round(2)

    merged = monthly_defect.merge(monthly_sales, on="sales_month", how="inner")
    merged = merged.sort_values("sales_month").reset_index(drop=True)
    merged["sales_month"] = merged["sales_month"].dt.strftime("%Y-%m")
    merged.insert(0, group_col, group_value)
    return merged


def get_market_impact_score_by_product(
    tables: Tables,
    min_months: int = MIN_MONTHS_FOR_CORRELATION,
) -> pd.DataFrame:
    """product_id별 월별 불량률-시장점유율 상관계수 기반 영향도 점수.

    영향도 점수 = |corr(defect_rate_pct, market_share_est_pct)| x 3개년 누적
    매출액(revenue_krw). 원래 PRD 초안은 operating_margin_pct를 기준 지표로
    잡았으나, 실제 dataset_1을 검증해보니 operating_margin_pct가
    dim_company_product.target_margin_rate로 고정돼 있어 100개 제품 전부
    월별 분산이 0이었다(상관계수가 항상 미정의) -- 그래서 실제로 월별 변동이
    있는 market_share_est_pct로 교체했다(사람 확인 후 결정, docs/prd.md 7절
    참고). operating_margin_pct는 get_monthly_defect_vs_sales 결과에는 계속
    남아 있어 참고 지표로는 볼 수 있다.

    매칭된 월 수가 min_months 미만이거나 불량률에 변동이 없어 상관계수가
    정의되지 않으면(NaN) correlation/impact_score를 None으로 남긴다 --
    docs/prd.md 7절 "표본 크기" 캐비어트를 코드로 반영한 것이다. 누적
    매출액은 매칭된 월뿐 아니라 fact_market_sales 전체 기간 기준이다.
    """
    monthly = get_monthly_defect_vs_sales(tables)
    revenue_by_product = tables.fact_market_sales.groupby("product_id")["revenue_krw"].sum()

    rows = []
    for product_id, group in monthly.groupby("product_id"):
        matched_months = len(group)
        correlation = None
        if matched_months >= min_months:
            # 불량률/시장점유율에 변동이 전혀 없는 제품(표본이 작을 때 흔함)은
            # 분산이 0이라 상관계수가 정의되지 않는다(NaN) -- numpy가 이때 내는
            # "invalid value encountered in divide" 경고는 예상된 코드 경로다.
            with np.errstate(invalid="ignore", divide="ignore"):
                raw_corr = group["defect_rate_pct"].corr(group["market_share_est_pct"])
            if not pd.isna(raw_corr):
                correlation = round(float(raw_corr), 4)
        cumulative_revenue = float(revenue_by_product.get(product_id, 0.0))
        impact_score = round(abs(correlation) * cumulative_revenue, 2) if correlation is not None else None
        rows.append({
            "product_id": product_id,
            "matched_months": matched_months,
            "correlation": correlation,
            "cumulative_revenue_krw": cumulative_revenue,
            "impact_score": impact_score,
        })
    result = pd.DataFrame(rows)
    # product_id만으로는 사람이 알아보기 어려워, dim_company_product에서
    # 제품명/회사명을 붙여 대시보드 표에 그대로 노출할 수 있게 한다.
    result = result.merge(
        tables.dim_company_product[["product_id", "model_name", "company_name"]],
        on="product_id",
        how="left",
    )
    return result.sort_values("impact_score", ascending=False, na_position="last").reset_index(drop=True)


def _rollup_impact_score(tables: Tables, group_col: str, min_months: int) -> pd.DataFrame:
    """product_id 단위 영향도 점수를 equipment_id/factory_code로 합산 롤업.

    해당 축을 거친 product_id들의 impact_score를 그대로 더한다(docs/prd.md
    4.2절) -- 여러 설비/공장을 거친 제품의 점수가 각 축에 중복 반영될 수
    있다는 단순화를 그대로 따른 것이다. impact_score가 None인(표본 부족)
    제품은 0으로 취급해 합산하되, scored_product_count로 몇 개 제품이 실제
    점수에 기여했는지 함께 노출한다.
    """
    product_scores = get_market_impact_score_by_product(tables, min_months=min_months)
    pairs = tables.fact_production_run[[group_col, "product_id"]].drop_duplicates()
    merged = pairs.merge(product_scores[["product_id", "impact_score"]], on="product_id", how="left")
    grouped = merged.groupby(group_col).agg(
        product_count=("product_id", "nunique"),
        scored_product_count=("impact_score", lambda s: int(s.notna().sum())),
        impact_score=("impact_score", lambda s: float(s.fillna(0.0).sum())),
    ).reset_index()

    # equipment_id/factory_code도 ID만으로는 알아보기 어려우니 dim_equipment에서
    # 설비명/공장명을 붙인다. factory_code는 다수 equipment_id에 걸쳐 있으므로
    # (factory_code, factory_name) 조합만 중복 제거해 1:1로 맞춘다.
    if group_col == "equipment_id":
        labels = tables.dim_equipment[["equipment_id", "equipment_name", "factory_name"]].drop_duplicates()
        grouped = grouped.merge(labels, on="equipment_id", how="left")
    elif group_col == "factory_code":
        labels = tables.dim_equipment[["factory_code", "factory_name"]].drop_duplicates()
        grouped = grouped.merge(labels, on="factory_code", how="left")

    return grouped.sort_values("impact_score", ascending=False).reset_index(drop=True)


def get_market_impact_score_by_equipment(
    tables: Tables,
    min_months: int = MIN_MONTHS_FOR_CORRELATION,
) -> pd.DataFrame:
    """equipment_id별 영향도 점수 롤업 -- 해당 설비를 거친 product_id들의 impact_score 합."""
    return _rollup_impact_score(tables, "equipment_id", min_months)


def get_market_impact_score_by_factory(
    tables: Tables,
    min_months: int = MIN_MONTHS_FOR_CORRELATION,
) -> pd.DataFrame:
    """factory_code별 영향도 점수 롤업."""
    return _rollup_impact_score(tables, "factory_code", min_months)


def get_revenue_projection(
    tables: Tables,
    entity_id: str,
    target_defect_rate_pct: Optional[float] = None,
    group_col: str = "product_id",
) -> dict:
    """"불량률을 낮추면 매출이 얼마나 바뀌는가"를 정량화하는 what-if 분석.

    get_monthly_defect_vs_sales_by_group의 (defect_rate_pct, revenue_krw) 월별
    쌍에 최소자승 직선(y = slope*x + intercept)을 적합해, 목표 불량률에서의
    예상 매출을 계산한다. get_market_impact_score_by_*의 상관계수(방향성만
    표시)와 달리 이 함수는 실제 기울기(원/%p)를 계산해 "1%p 낮추면 몇 원"을
    구체적으로 답한다.

    group_col="product_id"(기본)면 entity_id는 product_id다. "equipment_id"/
    "factory_code"를 넘기면 해당 설비/공장을 거친 제품군의 매출 흐름을
    근사해서 사용한다 -- get_monthly_defect_vs_sales_by_group의 단순화가
    그대로 적용된다(특정 월의 실제 생산 여부는 따지지 않음).

    target_defect_rate_pct를 지정하지 않으면 실제로 달성했던 최저 불량률
    (과거 최고 품질 달성 월)을 목표로 삼는다 -- 임의의 목표가 아니라
    데이터에 근거한 현실적인 시나리오다.

    한계(docs/prd.md 7절과 동일): 상관관계이지 인과관계가 아니다. 회귀선은
    관측된 불량률 범위 밖으로 외삽할수록(특히 matched_months가 적을수록)
    신뢰도가 낮아진다 -- 그래서 r_squared와 matched_months를 함께 반환해
    호출측(FE)이 신뢰도를 같이 보여주게 한다.
    """
    monthly = get_monthly_defect_vs_sales_by_group(tables, group_col, entity_id)
    n = len(monthly)

    result: dict = {
        group_col: entity_id,
        "matched_months": n,
        "slope_krw_per_pp": None,
        "intercept_krw": None,
        "r_squared": None,
        "current_avg_defect_rate_pct": None,
        "current_avg_revenue_krw": None,
        "target_defect_rate_pct": target_defect_rate_pct,
        "projected_revenue_krw": None,
        "revenue_delta_krw": None,
        "revenue_delta_pct": None,
    }
    if n == 0:
        return result

    x = monthly["defect_rate_pct"].to_numpy(dtype=float)
    y = monthly["revenue_krw"].to_numpy(dtype=float)
    current_avg_defect_rate = float(x.mean())
    current_avg_revenue = float(y.mean())
    result["current_avg_defect_rate_pct"] = round(current_avg_defect_rate, 2)
    result["current_avg_revenue_krw"] = round(current_avg_revenue, 2)

    if target_defect_rate_pct is None:
        target_defect_rate_pct = float(x.min())
        result["target_defect_rate_pct"] = round(target_defect_rate_pct, 2)

    # 불량률 변동이 전혀 없으면(n<2 포함) 직선을 적합할 수 없다 -- 회귀 관련
    # 필드는 None으로 남기고 여기서 끝낸다.
    if n < 2 or np.allclose(x, x[0]):
        return result

    with np.errstate(invalid="ignore", divide="ignore"):
        slope, intercept = np.polyfit(x, y, 1)
        y_pred = slope * x + intercept
        ss_res = float(np.sum((y - y_pred) ** 2))
        ss_tot = float(np.sum((y - y.mean()) ** 2))
    r_squared = (1 - ss_res / ss_tot) if ss_tot > 0 else None

    projected_revenue = float(slope * target_defect_rate_pct + intercept)
    delta = projected_revenue - current_avg_revenue

    result["slope_krw_per_pp"] = round(float(slope), 2)
    result["intercept_krw"] = round(float(intercept), 2)
    result["r_squared"] = round(float(r_squared), 4) if r_squared is not None else None
    result["projected_revenue_krw"] = round(projected_revenue, 2)
    result["revenue_delta_krw"] = round(delta, 2)
    result["revenue_delta_pct"] = (
        round(delta / current_avg_revenue * 100, 2) if current_avg_revenue else None
    )
    return result
